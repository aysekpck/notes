#openpyxl library
#cmd komut satırını açın
#pip install openpyxl  yazın
from openpyxl import *  #modül ön eki kullanmadan özelliklere erişim
#kütüphaneyi programa dahil ediyoruz
#workbook():dosya oluşururken ilk  olarak excel çalışma kitabı tanımlanır.
kitap=Workbook()
sheet=kitap.active
kitap.save("nurcan.xlsx")
kitap.close()
#worksheet işlemleri
#pythonda bir excel kitaplğı içerisinde
#yeni bir çalışma sayfası oluşturulabilir
kitap.create_sheet("calıismasayfasi1")
kitap.create_sheet("calıismasayfasi2",2)
#DOSYAYA İÇERİK EKLEMEK
kitap=Workbook()
sheet=kitap.active
sheet.append(("YbS","Bilgisayar","","Dogus")) #demet
sheet.append(["python","veri","bilim"]) #liste
kitap.save("nurcan.xlsx")
kitap.close()
#HÜCRELERE VERİ YAZDIRMAK
kitap=Workbook()
sheet=kitap.active
sheet["E4"]="KORONA"
sheet["D6"]="bitsin"
kitap.save("nurcan.xlsx")
kitap.close()

#2.yol:
kitap=Workbook()
sheet=kitap.active
sheet.cell(row=3,column=4,value="Excele veri yuklemeyi öğren")
kitap.save("sefa.xlsx")
kitap.close()

#Varolan bir Excel Dosyasını Açmak ve Düzenlemek
#load_workbook("üzerinde işlem yapılacak dosya") fonksiyonu kullanılır
kitap=load_workbook("carsamba.xlsx")
sheet=kitap.create_sheet("calismasayfası",1)
#calisma sayfası adında sayfa açıyoruz excelde
sheet.cell(row=12,column=4,value="DOGUS UNIVERSITY")
kitap.save("carsamba.xlsx")
kitap.close()

#Varolan sayfadan veri okumak
kitap=load_workbook("carsamba.xlsx")
sheet=kitap.active
d1=sheet["A6"]
d2=sheet["B6"]
d3=sheet.cell(9,3)
print(d1.value)
print(d2.value)
print(d3.value)

kitap.close()
#satır satır veri okumak

kitap=load_workbook("carsamba.xlsx")
sheet=kitap.active
#iter_rows() X  iter_columns()
for row in sheet.iter_rows(min_row=1,min_col=1,max_row=9,max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()
kitap.close()

#sütun sütun veri okumak için
kitap=load_workbook("carsamba.xlsx")
sheet=kitap.active
#iter_rows() X  iter_columns()
for row in sheet.iter_cols(min_row=1,min_col=1,max_row=9,max_col=3):
    for cell in row:
        print(cell.value)
    print()
kitap.close()












